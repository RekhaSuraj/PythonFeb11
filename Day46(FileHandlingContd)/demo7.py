# read from an excel file
import openpyxl

wb = openpyxl.load_workbook('Book1.xlsx')
sheet_name = wb['Sheet1']
# print(sheet_name.cell(2,3).value)
# c2

# print(sheet_name.cell(1,2).value)
# test1

# to print 4 rows and 3 columns
for row in sheet_name.iter_rows(min_row=1,max_row=4,min_col=1,max_col=3):
    for cell in row:
        print(cell.value,end='\t') #tab between the values for display purpose
    print()

# 1	test1	c1
# 2	test2	c2
# 3	test3	c3
# 4	test4	c4
